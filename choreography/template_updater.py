import os
import csv
from openpyxl import load_workbook

script_directory = os.path.dirname(os.path.abspath(__file__))

# Function to get list of files in 'moves' folder
def get_move_files():
    move_files = {}
    moves_folder = os.path.join(script_directory, "moves")
    if os.path.exists(moves_folder) and os.path.isdir(moves_folder):
        for file_name in os.listdir(moves_folder):
            file_path = os.path.join(moves_folder, file_name)
            if os.path.isfile(file_path):
                with open(file_path, 'r', newline='') as file:
                    csv_reader = csv.reader(file)
                    # Skip to row 2
                    next(csv_reader)
                    # Read duration value from column F (index 5)
                    for row in csv_reader:
                        if len(row) > 5:
                            move_files[file_name] = row[5]
                        else:
                            move_files[file_name] = ""
                        break  # Only read from the first row
    return move_files

# Function to modify template.ods
def modify_template():
    template_file = os.path.join(script_directory, "routines/template.xlsx")
    if os.path.exists(template_file):
        # Load the template file using openpyxl
        wb = load_workbook(template_file)
        
        # Get the 'movelist' sheet
        movelist_sheet = wb['movelist']

        # Clear the contents of 'movelist' sheet starting from row 2
        for row in movelist_sheet.iter_rows(min_row=2, max_col=movelist_sheet.max_column, max_row=movelist_sheet.max_row):
            for cell in row:
                cell.value = None

        # Get move files and their corresponding values
        move_files = get_move_files()

        # Add entries to movelist sheet starting from row 2
        row_index = 2
        for move_file, value in move_files.items():
            movelist_sheet.cell(row=row_index, column=1, value=move_file)
            movelist_sheet.cell(row=row_index, column=2, value=value)
            row_index += 1

        # Save modified data back to template.ods
        wb.save(template_file)
        print("Template modified successfully.")
    else:
        print("Template file not found.")

modify_template()
